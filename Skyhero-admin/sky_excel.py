from openpyxl import load_workbook
import PySimpleGUI as sg
import datetime

date_object = datetime.date.today()
choices=("Closed", "Service Activity", "In progress", "No Answer", "Left Message")
lunchChoices=("Daily Calls", "Chat session", "Moderation", "Xfr Calls MM", "Req to check case in the morning",
	"Complaints from MM/Feed Back", "Good Feed back from MM & Client", "All Internal Complaints", "Late mark, Half day & Leave",
	"Helping UK/SA Engineer")

window=sg.Window("Sky-excel",
	layout=[
	[sg.Text("Choose file:"),sg.In(enable_events=True, justification='center',key="YourSheet"), sg.FileBrowse()],
	[sg.Text("Ticket Id:   "),sg.InputText(justification='center',key="tid")],
	[sg.Text("Start Time: "),sg.InputText(justification='center',key="st")],
	[sg.Text("End Time:  "),sg.InputText(justification='center',key="et")],
	[sg.Text("Status:         "),sg.Listbox(choices,size=(25,len(choices)),key="stats")],
	[sg.Text("Lunch Time:  "),sg.Listbox(lunchChoices,size=(25,len(choices)),key="lunchStat")],
	[sg.Text("Company name:"),sg.InputText(justification='center',key="cn")],
	[sg.Text("Issue title:         "),sg.InputText(justification='center',key="it")],
	[sg.Text("Comments:    "), sg.Multiline(size=(50,7),justification='center',key="com")],
	[sg.Button("OK")]
	],
	margins=(100,50))
i=3
while True:
	event, values=window.read()
	if event=="Exit" or event==sg.WIN_CLOSED:
		break
	if event=="OK":
		if values["YourSheet"]:
			print(values["YourSheet"])
			wb = load_workbook(values["YourSheet"])
			ws = wb.active
			f = open(values["YourSheet"],"r")
		
		while True:
			my_file = open(f"{date_object}.txt","a+")
			if ws[f"A{i}"].value!=None:
				i+=1
			elif (ws[f"A{i}"].value)==None:
				ws[f"A{i}"]=str(values["tid"])
				ws[f"C{i}"]=str(values["st"])
				ws[f"D{i}"]=str(values["et"])
				ws[f"F{i}"]=str(values["stats"][0])
				ws[f"G{i}"]=str(values["lunchStat"][0])
				wb.save(values["YourSheet"])

				my_file.write(str(values["st"])+"\n")
				my_file.write(str(values["et"])+"\n")
				my_file.write(str(values["cn"])+"\n")
				my_file.write(str(values["tid"])+"\n")
				my_file.write(str(values["stats"][0])+"\n")
				my_file.write(str(values["it"])+"\n\n")
				my_file.write(str(values["com"])+"\n")
				my_file.write("---------------\n")

				break
		my_file.close()
		window["st"]("")
		window["et"]("")
		window["cn"]("")
		window["tid"]("")
		window["it"]("")
		window["com"]("")

		print("Done")
